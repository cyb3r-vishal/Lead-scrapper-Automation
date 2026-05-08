"""Incremental Excel writer — survives crashes mid-run."""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

log = logging.getLogger("excel")


class ExcelWriter:
    """Appends lead rows to an .xlsx file. Flushes on every append so a crash
    never loses rows already extracted."""

    def __init__(self, path: str, columns: list[str]) -> None:
        self.path = Path(path)
        # Always track where each lead came from + when, regardless of AI schema
        meta = ["source_url", "scraped_at"]
        self.columns = list(dict.fromkeys([*columns, *meta]))  # dedupe, preserve order
        self._init_file()

    def _init_file(self) -> None:
        if self.path.exists():
            wb = load_workbook(self.path)
            ws = wb.active
            existing = [c.value for c in ws[1]] if ws.max_row >= 1 else []
            if existing != self.columns:
                # Schema mismatch — back up and start fresh so we don't corrupt data.
                backup = self.path.with_suffix(
                    f".bak-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
                )
                self.path.rename(backup)
                log.warning("Schema changed; backed up prior file to %s", backup)
                self._create_new()
            # else: file is valid, keep appending
        else:
            self._create_new()

    def _create_new(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(self.columns)
        wb.save(self.path)

    def append(self, row: dict[str, Any]) -> None:
        row = {**row, "scraped_at": datetime.now().isoformat(timespec="seconds")}
        wb = load_workbook(self.path)
        ws = wb.active
        ws.append([self._stringify(row.get(c, "")) for c in self.columns])
        wb.save(self.path)

    @staticmethod
    def _stringify(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, (list, tuple, set)):
            return ", ".join(str(x) for x in v)
        return str(v)

    def row_count(self) -> int:
        wb = load_workbook(self.path)
        return max(wb.active.max_row - 1, 0)  # minus header

    def existing_rows(self) -> list[dict[str, Any]]:
        """Read back all rows already in the file, as dicts keyed by column.
        Used at startup so a re-run against an existing file won't add dupes."""
        wb = load_workbook(self.path)
        ws = wb.active
        rows: list[dict[str, Any]] = []
        header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if not header:
            return rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({h: v for h, v in zip(header, row) if h})
        return rows
